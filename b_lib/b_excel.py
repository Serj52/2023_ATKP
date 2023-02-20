import pandas
import openpyxl
import logging
import keyring
from b_lib.EXCEPTION_HANDLER import ExctractPWDError


class Excel:
    # @staticmethod
    # def get_pwd(organization, file):
    #TODO:удалить после тестов
    #     try:
    #         workbook = openpyxl.load_workbook(file)
    #         worksheet = workbook.active
    #         max_row = worksheet.max_row
    #         pwd = ''
    #         for row in range(2, max_row + 1):
    #             value = worksheet.cell(row=row, column=1).value
    #             if value == organization:
    #                 username = worksheet.cell(row=row, column=2).value
    #                 pwd = keyring.get_password('eosdo', username)
    #                 logging.info(f'Пароль для {organization} извлечен')
    #                 workbook.close()
    #                 return {'pwd': pwd, 'username': username}
    #         if pwd == '':
    #             logging.error(f'Проверьте наличие username для {organization} в {file}')
    #             workbook.close()
    #             raise ExctractPWDError('Ошибка при получении пароля ЕОСДО')
    #     except Exception as err:
    #         logging.error(err)
    #         raise ExctractPWDError('Ошибка при получении пароля ЕОСДО')

    @staticmethod
    def get_pwd(organization, path):
        dataframe = pandas.read_excel(path, dtype=dict)
        organization = dataframe.loc[dataframe['Предприятие'] == f'{organization}']
        if organization.empty:
            raise ExctractPWDError('Ошибка при получении пароля ЕОСДО')
        else:
            username = organization['УЗ'].values[0]
            if pandas.isna(username):
                raise ExctractPWDError('Ошибка при получении пароля ЕОСДО')
            else:
                pwd = keyring.get_password('eosdo', username)
                return {'pwd': pwd, 'username': username}

    def get_organization(self, file):
        workbook = openpyxl.load_workbook(file)
        worksheet = workbook.active
        max_row = worksheet.max_row
        organizations = []
        for row in range(2, max_row + 1):
            value = worksheet.cell(row=row, column=1).value
            if value is not None:
                organizations.append(value)
        return organizations


if __name__ == '__main__':
    excel = Excel()
    excel.get_pwd('АО "Гринатом"')